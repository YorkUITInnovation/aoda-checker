"""
Excel report generation for accessibility scans.
Exports page scan data to XLSX format.
"""
import io
from datetime import datetime
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from src.models import ScanResult, PageResult


def generate_excel_report(scan_result: ScanResult) -> io.BytesIO:
    """
    Generate an Excel (XLSX) report of the scan results.
    
    Args:
        scan_result: The scan result to export
        
    Returns:
        BytesIO object containing the Excel file
    """
    # Create workbook
    wb = Workbook()
    
    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_summary_sheet(ws_summary, scan_result)
    
    # Create Page Details sheet
    ws_pages = wb.create_sheet("Page Details")
    _create_page_details_sheet(ws_pages, scan_result)
    
    # Create Violations sheet
    ws_violations = wb.create_sheet("All Violations")
    _create_violations_sheet(ws_violations, scan_result)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output


def _create_summary_sheet(ws, scan_result: ScanResult):
    """Create the summary sheet with scan overview."""
    # Header style
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    
    # Title
    ws['A1'] = "Accessibility Scan Summary"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')
    
    # Scan Information
    row = 3
    info_items = [
        ("Scan ID:", scan_result.scan_id),
        ("Start URL:", scan_result.start_url),
        ("Scan Date:", scan_result.start_time.strftime('%Y-%m-%d %H:%M:%S')),
        ("Duration:", f"{scan_result.duration:.2f} seconds" if scan_result.duration else "N/A"),
        ("Scan Mode:", scan_result.scan_mode.upper()),
        ("", ""),
        ("Pages Scanned:", scan_result.pages_scanned),
        ("Pages with Violations:", scan_result.pages_with_violations),
        ("Total Violations:", scan_result.total_violations),
    ]
    
    for label, value in info_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Violations by Severity
    row += 1
    ws[f'A{row}'] = "Violations by Severity"
    ws[f'A{row}'].font = Font(bold=True, size=12)
    ws.merge_cells(f'A{row}:B{row}')
    row += 1
    
    violations_by_severity = scan_result.get_violations_by_severity()
    severity_items = [
        ("Errors:", violations_by_severity.get('error', 0)),
        ("Warnings:", violations_by_severity.get('warning', 0)),
        ("Alerts:", violations_by_severity.get('alert', 0)),
    ]
    
    for label, value in severity_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        ws[f'A{row}'].font = Font(bold=True)
        row += 1
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 60


def _create_page_details_sheet(ws, scan_result: ScanResult):
    """Create the page details sheet with all scanned pages."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Page URL", "Violations", "Errors", "Warnings", "Alerts", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 2
    for page in scan_result.page_results:
        # Count violations by severity for this page
        error_count = sum(1 for v in page.violations if v.severity == 'error')
        warning_count = sum(1 for v in page.violations if v.severity == 'warning')
        alert_count = sum(1 for v in page.violations if v.severity == 'alert')
        unknown_count = sum(1 for v in page.violations if not v.severity or v.severity not in ['error', 'warning', 'alert'])

        total_violations = len(page.violations)

        ws.cell(row=row, column=1, value=page.url).border = border
        ws.cell(row=row, column=2, value=total_violations).border = border
        ws.cell(row=row, column=3, value=error_count).border = border
        ws.cell(row=row, column=4, value=warning_count).border = border
        ws.cell(row=row, column=5, value=alert_count).border = border
        ws.cell(row=row, column=6, value="Pass" if total_violations == 0 else "Issues Found").border = border

        # Color code the status
        status_cell = ws.cell(row=row, column=6)
        if total_violations == 0:
            status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        elif error_count > 0:
            status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        elif warning_count > 0:
            status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        else:
            status_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        row += 1
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 80
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 15


def _create_violations_sheet(ws, scan_result: ScanResult):
    """Create the violations sheet with all violations."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ["Page URL", "Severity", "Issue Type", "Description", "Help URL", "Tags"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    
    # Data rows
    row = 2
    for page in scan_result.page_results:
        for violation in page.violations:
            # Handle None values
            severity = violation.severity or violation.impact.value if violation.impact else 'unknown'
            severity_text = severity.upper() if severity else 'UNKNOWN'

            # Get tags as comma-separated string
            tags_str = ", ".join(violation.tags) if violation.tags else "N/A"

            ws.cell(row=row, column=1, value=page.url).border = border
            ws.cell(row=row, column=2, value=severity_text).border = border
            ws.cell(row=row, column=3, value=violation.id or 'N/A').border = border
            ws.cell(row=row, column=4, value=violation.description or violation.help or 'N/A').border = border
            ws.cell(row=row, column=5, value=violation.help_url or "N/A").border = border
            ws.cell(row=row, column=6, value=tags_str).border = border

            # Color code by severity
            severity_cell = ws.cell(row=row, column=2)
            if severity == 'error':
                severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif severity == 'warning':
                severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            elif severity == 'alert':
                severity_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
            else:
                # Unknown or None severity - gray background
                severity_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

            row += 1
    
    # Auto-size columns
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 50
    ws.column_dimensions['E'].width = 40
    ws.column_dimensions['F'].width = 30


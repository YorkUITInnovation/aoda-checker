"""
Bulk Excel report generation for multiple accessibility scans.
Exports combined data from multiple scans to XLSX format.
"""
import io
from typing import List
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from src.models import ScanResult


def generate_bulk_excel_report(scan_results: List[ScanResult]) -> io.BytesIO:
    """
    Generate an Excel (XLSX) report combining multiple scan results.

    Args:
        scan_results: List of scan results to combine

    Returns:
        BytesIO object containing the Excel file
    """
    # Create workbook
    wb = Workbook()

    # Create Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _create_combined_summary_sheet(ws_summary, scan_results)

    # Create Page Details sheet
    ws_pages = wb.create_sheet("Page Details")
    _create_combined_page_details_sheet(ws_pages, scan_results)

    # Create Violations sheet
    ws_violations = wb.create_sheet("All Violations")
    _create_combined_violations_sheet(ws_violations, scan_results)

    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return output


def _create_combined_summary_sheet(ws, scan_results: List[ScanResult]):
    """Create the summary sheet with combined scan overview."""
    # Header style
    header_font = Font(bold=True, size=14, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Title
    ws['A1'] = f"Combined Accessibility Scan Summary ({len(scan_results)} Scans)"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:B1')

    # Overall statistics
    total_pages = sum(s.pages_scanned for s in scan_results)
    total_pages_with_violations = sum(s.pages_with_violations for s in scan_results)
    total_violations = sum(s.total_violations for s in scan_results)

    # Calculate combined violations by severity
    combined_errors = 0
    combined_warnings = 0
    combined_alerts = 0

    for scan in scan_results:
        violations_by_severity = scan.get_violations_by_severity()
        combined_errors += violations_by_severity.get('error', 0)
        combined_warnings += violations_by_severity.get('warning', 0)
        combined_alerts += violations_by_severity.get('alert', 0)

    # Overall Summary
    row = 3
    ws[f'A{row}'] = "Overall Statistics"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    overall_items = [
        ("Total Scans:", len(scan_results)),
        ("Total Pages Scanned:", total_pages),
        ("Pages with Violations:", total_pages_with_violations),
        ("Total Violations:", total_violations),
        ("", ""),
        ("Errors:", combined_errors),
        ("Warnings:", combined_warnings),
        ("Alerts:", combined_alerts),
    ]

    for label, value in overall_items:
        ws[f'A{row}'] = label
        ws[f'B{row}'] = value
        if label:
            ws[f'A{row}'].font = Font(bold=True)
        row += 1

    # Individual Scans
    row += 1
    ws[f'A{row}'] = "Individual Scans"
    ws[f'A{row}'].font = Font(bold=True, size=12, color="FFFFFF")
    ws[f'A{row}'].fill = header_fill
    ws[f'B{row}'].fill = header_fill
    ws.merge_cells(f'A{row}:B{row}')
    row += 1

    # Table header
    ws[f'A{row}'] = "URL"
    ws[f'B{row}'] = "Date"
    ws[f'C{row}'] = "Pages"
    ws[f'D{row}'] = "Violations"
    ws[f'E{row}'] = "Errors"
    ws[f'F{row}'] = "Warnings"
    ws[f'G{row}'] = "Alerts"

    for col in range(1, 8):
        cell = ws.cell(row=row, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border
    row += 1

    # Individual scan data
    for scan in scan_results:
        violations_by_severity = scan.get_violations_by_severity()

        ws.cell(row=row, column=1, value=scan.start_url).border = border
        ws.cell(row=row, column=2, value=scan.start_time.strftime('%Y-%m-%d %H:%M')).border = border
        ws.cell(row=row, column=3, value=scan.pages_scanned).border = border
        ws.cell(row=row, column=4, value=scan.total_violations).border = border
        ws.cell(row=row, column=5, value=violations_by_severity.get('error', 0)).border = border
        ws.cell(row=row, column=6, value=violations_by_severity.get('warning', 0)).border = border
        ws.cell(row=row, column=7, value=violations_by_severity.get('alert', 0)).border = border

        row += 1

    # Auto-size columns
    ws.column_dimensions['A'].width = 60
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 12
    ws.column_dimensions['E'].width = 10
    ws.column_dimensions['F'].width = 12
    ws.column_dimensions['G'].width = 10


def _create_combined_page_details_sheet(ws, scan_results: List[ScanResult]):
    """Create the page details sheet with all scanned pages from all scans."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Scan URL", "Page URL", "Violations", "Errors", "Warnings", "Alerts", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    row = 2
    for scan in scan_results:
        for page in scan.page_results:
            # Count violations by severity for this page
            error_count = sum(1 for v in page.violations if v.severity == 'error')
            warning_count = sum(1 for v in page.violations if v.severity == 'warning')
            alert_count = sum(1 for v in page.violations if v.severity == 'alert')
            total_violations = len(page.violations)

            ws.cell(row=row, column=1, value=scan.start_url).border = border
            ws.cell(row=row, column=2, value=page.url).border = border
            ws.cell(row=row, column=3, value=total_violations).border = border
            ws.cell(row=row, column=4, value=error_count).border = border
            ws.cell(row=row, column=5, value=warning_count).border = border
            ws.cell(row=row, column=6, value=alert_count).border = border
            ws.cell(row=row, column=7, value="Pass" if total_violations == 0 else "Issues Found").border = border

            # Color code the status
            status_cell = ws.cell(row=row, column=7)
            if total_violations == 0:
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            elif error_count > 0:
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            elif warning_count > 0:
                status_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            else:
                status_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")

            row += 1

    # Auto-size columns
    ws.column_dimensions['A'].width = 50
    ws.column_dimensions['B'].width = 50
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 10
    ws.column_dimensions['E'].width = 12
    ws.column_dimensions['F'].width = 10
    ws.column_dimensions['G'].width = 15


def _create_combined_violations_sheet(ws, scan_results: List[ScanResult]):
    """Create the violations sheet with all violations from all scans."""
    # Header style
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="0066CC", end_color="0066CC", fill_type="solid")
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Headers
    headers = ["Scan URL", "Page URL", "Severity", "Issue Type", "Description", "Help URL", "Tags"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        cell.border = border

    # Data rows
    row = 2
    for scan in scan_results:
        for page in scan.page_results:
            for violation in page.violations:
                # Handle None values
                severity = violation.severity or violation.impact.value if violation.impact else 'unknown'
                severity_text = severity.upper() if severity else 'UNKNOWN'

                # Get tags as comma-separated string
                tags_str = ", ".join(violation.tags) if violation.tags else "N/A"

                ws.cell(row=row, column=1, value=scan.start_url).border = border
                ws.cell(row=row, column=2, value=page.url).border = border
                ws.cell(row=row, column=3, value=severity_text).border = border
                ws.cell(row=row, column=4, value=violation.id or 'N/A').border = border
                ws.cell(row=row, column=5, value=violation.description or violation.help or 'N/A').border = border
                ws.cell(row=row, column=6, value=violation.help_url or "N/A").border = border
                ws.cell(row=row, column=7, value=tags_str).border = border

                # Color code by severity
                severity_cell = ws.cell(row=row, column=3)
                if severity == 'error':
                    severity_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                elif severity == 'warning':
                    severity_cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
                elif severity == 'alert':
                    severity_cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
                else:
                    # Unknown or None severity - gray background
                    severity_cell.fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")

                row += 1

    # Auto-size columns
    ws.column_dimensions['A'].width = 40
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 30
    ws.column_dimensions['E'].width = 50
    ws.column_dimensions['F'].width = 40
    ws.column_dimensions['G'].width = 30

